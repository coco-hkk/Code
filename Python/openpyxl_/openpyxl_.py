"""Excel 操作用例

openpyxl 3.0.10

作者：coco-hkk
日期：2022年9月9日
"""
import openpyxl
from openpyxl.styles import PatternFill

class ReadExcel:
    """操作 Excel 文件"""

    def __init__(self, file_name, sheet_name):
        """初始化方法

        Args:
            fileName: Excel 文件名
            sheetName: 表单名
        """
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb_ = None
        self.ws_ = None

    def open(self):
        """打开工作簿，选中表单

        访问 worksheets 可使用索引或表名访问，如访问第一个表
        self.wb_.worksheets[0] 或 self.wb_["Sheet1"]
        """
        self.wb_ = openpyxl.load_workbook(self.file_name)
        if isinstance(self.sheet_name, int):
            self.ws_ = self.wb_.worksheets[self.sheet_name]
        else:
            self.ws_ = self.wb_[self.sheet_name]

    def save(self):
        """保存工作簿，关闭 Excel"""
        self.wb_.save(self.file_name)
        self.wb_.close()

    def read_data_dict(self):
        """读取数据，将每条数据存储为字典类型

        Returns:
            存储每行数据的列表，每行数据以字典方式存储
        """
        self.open()

        # 通过 rows 获取 Excel 文件中所有的行数据，然后把数据转换成列表
        rows = list(self.ws_.rows)

        # 表头
        title = []
        # 遍历 Excel 文件中的第一行表头信息，依据不同表的表头所在行更改值
        # 从 0 开始计数，0 表示第一行
        for i in rows[4]:
            title.append(i.value)

        # 数据列表
        datas = []
        # 遍历数据行，从第六行开始到倒数第四行为有效数据范围
        for row in rows[5:-3]:
            # 定义一个列表存放每行数据
            data = []
            for r__ in row:
                data.append(r__.value)

            # 将每行的数据通过 zip 进行打包，然后转成字典
            datas.append(dict(zip(title, data)))

        return datas

    def write_data(self, row, column, value):
        """写入数据

        注意 Excel 表格不能为只读属性，程序运行时不要手动打开 Excel，否则提示
        Permission Denied.

        Args:
            row: 行。第一行从 1 开始计数
            column: 列。同上，如 A 列，用 1 表示
            value: 数据
        """
        self.open()

        # 指定行列进行数据入
        self.ws_.cell(row=row, column=column, value=value)

        self.save()

    def cell_fillcolor_set(self, row, column, fillcolor):
        """设置单元格的填充色

        Args:
            row: 行
            column: 列
            fillcolor: 填充色
        """

        self.open()

        # 填充色
        fills = PatternFill("solid", fgColor=fillcolor)

        # 指定行列进行数据入
        self.ws_.cell(row=row, column=column).fill = fills

        self.save()

if __name__ == "__main__":
    xlsx = ReadExcel("your.xlsx", "Sheet1")
    xlsx_data = xlsx.read_data_dict()

    print(xlsx_data)
    xlsx.write_data(6,1,"test")
    xlsx.cell_fillcolor_set(6,1,"FF0000")
